import json
from io import BytesIO
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from llm_client import get_openai


SENTIMENT_ENUM = ["Positive", "Negative", "Neutral"]


def compute_aht_seconds(transcript_df: pd.DataFrame) -> float:
    """AHT = (max END_TIME_MS - min START_TIME_MS) / 1000."""
    try:
        start = float(transcript_df["START_TIME_MS"].min())
        end = float(transcript_df["END_TIME_MS"].max())
        return max(0.0, (end - start) / 1000.0)
    except Exception:
        return 0.0


def transcript_to_one_cell(transcript_df: pd.DataFrame) -> str:
    """Join transcript into one cell: 'speaker: text' lines."""
    df = transcript_df.sort_values("phrase_rank")
    lines = []
    for r in df.to_dict("records"):
        spk = str(r.get("speaker", "")).strip()
        txt = str(r.get("text", "")).strip()
        if txt:
            lines.append(f"{spk}: {txt}" if spk else txt)
    return "\n".join(lines).strip()


def compact_text_for_llm(full_text: str, max_chars: int = 8000) -> str:
    """Trim large transcripts for token safety while keeping key parts."""
    if not full_text:
        return ""
    if len(full_text) <= max_chars:
        return full_text
    head = full_text[: int(max_chars * 0.65)]
    tail = full_text[-int(max_chars * 0.35) :]
    return head + "\n...\n" + tail


# ---------------------------
# OPEN-ENDED: Discover tree
# ---------------------------

def _build_discovery_messages(sample_payload: List[dict]) -> List[dict]:
    system = {
        "role": "system",
        "content": (
            "You are an expert intent taxonomist for telecom/call-center transcripts. "
            "You will infer a practical two-level intent hierarchy (L1, L2) that can classify a whole dataset."
        ),
    }
    user = {
        "role": "user",
        "content": (
            "<TASK>\n"
            "Given transcript samples, create an intent hierarchy with:\n"
            "- L1: broad intent categories (6–12 max)\n"
            "- L2: specific sub-intents under each L1 (4–12 each)\n"
            "The hierarchy should be useful for classifying the entire dataset.\n"
            "Use concise labels. Avoid duplicates.\n"
            "</TASK>\n\n"
            "<SAMPLES>\n"
            f"{json.dumps(sample_payload)}\n"
            "</SAMPLES>\n\n"
            "<OUTPUT>\n"
            "Return JSON:\n"
            "{\n"
            '  "intent_tree": [\n'
            '    {"L1": "string", "L2": ["string", "..."]},\n'
            "    ...\n"
            "  ]\n"
            "}\n"
            "</OUTPUT>"
        ),
    }
    return [system, user]


def discover_intent_tree(transcripts_text: List[str]) -> Dict[str, List[str]]:
    """
    Build a dataset-wide intent tree from transcript samples.
    Returns dict: {L1: [L2,...]}.
    """
    client, model = get_openai()

    # Make compact sample payload
    sample_payload = []
    for i, t in enumerate(transcripts_text, start=1):
        sample_payload.append({"sample_id": i, "transcript": compact_text_for_llm(t, max_chars=4000)})

    schema = {
        "name": "intent_tree_discovery",
        "strict": True,
        "schema": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "intent_tree": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "additionalProperties": False,
                        "properties": {
                            "L1": {"type": "string"},
                            "L2": {"type": "array", "items": {"type": "string"}},
                        },
                        "required": ["L1", "L2"],
                    },
                }
            },
            "required": ["intent_tree"],
        },
    }

    resp = client.chat.completions.create(
        model=model,
        messages=_build_discovery_messages(sample_payload),
        response_format={"type": "json_schema", "json_schema": schema},
    )
    parsed = json.loads(resp.choices[0].message.content)

    tree: Dict[str, List[str]] = {}
    for node in parsed["intent_tree"]:
        l1 = str(node["L1"]).strip()
        l2s = [str(x).strip() for x in node.get("L2", []) if str(x).strip()]
        # Ensure at least "Other"
        if "Other" not in l2s:
            l2s.append("Other")
        if l1:
            tree[l1] = sorted(list(dict.fromkeys(l2s)))
    return tree


# ---------------------------
# Classify one transcript
# ---------------------------

def _build_mapping_messages(
    transcript_text: str,
    intent_tree: Dict[str, List[str]],
    mode: str,
) -> List[dict]:
    """
    mode:
      - "open-ended": tree is discovered (still constrained)
      - "close-ended": tree is uploaded/final (strict mapping)
    """
    system = {
        "role": "system",
        "content": (
            "You are an expert call-center intent classifier. "
            "You must map each transcript to a primary intent at two levels (L1, L2) and sentiment."
        ),
    }

    user = {
        "role": "user",
        "content": (
            "<RULES>\n"
            "1) Assign ONLY ONE primary intent (even if multiple are present).\n"
            "2) L1 must be one of the provided L1 values.\n"
            "3) L2 must be one of the provided L2 values under the chosen L1.\n"
            "4) Sentiment must be exactly: Positive, Negative, or Neutral.\n"
            "5) Base sentiment on the overall customer tone/outcome.\n"
            "</RULES>\n\n"
            f"<MODE>\n{mode}\n</MODE>\n\n"
            "<INTENT_TREE>\n"
            f"{json.dumps(intent_tree)}\n"
            "</INTENT_TREE>\n\n"
            "<TRANSCRIPT>\n"
            f"{compact_text_for_llm(transcript_text, max_chars=8000)}\n"
            "</TRANSCRIPT>\n\n"
            "<OUTPUT>\n"
            "Return JSON:\n"
            "{\n"
            '  "L1": "string",\n'
            '  "L2": "string",\n'
            '  "Sentiment": "Positive|Negative|Neutral"\n'
            "}\n"
            "</OUTPUT>"
        ),
    }
    return [system, user]


def classify_transcript_intent(
    transcript_text: str,
    intent_tree: Dict[str, List[str]],
    mode: str,
) -> Tuple[str, str, str]:
    """
    Returns (L1, L2, Sentiment).
    Validates L2 belongs to L1, else sets L2="Other" (if available).
    """
    client, model = get_openai()

    l1_enum = sorted(intent_tree.keys())
    l2_enum = sorted({l2 for l2s in intent_tree.values() for l2 in l2s} | {"Other"})

    schema = {
        "name": "intent_mapping",
        "strict": True,
        "schema": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "L1": {"type": "string", "enum": l1_enum},
                "L2": {"type": "string", "enum": l2_enum},
                "Sentiment": {"type": "string", "enum": SENTIMENT_ENUM},
            },
            "required": ["L1", "L2", "Sentiment"],
        },
    }

    resp = client.chat.completions.create(
        model=model,
        messages=_build_mapping_messages(transcript_text, intent_tree, mode),
        response_format={"type": "json_schema", "json_schema": schema},
    )
    parsed = json.loads(resp.choices[0].message.content)

    L1 = str(parsed["L1"]).strip()
    L2 = str(parsed["L2"]).strip()
    Sent = str(parsed["Sentiment"]).strip()

    # Validate hierarchy
    if L1 not in intent_tree:
        # Should not happen due to enum, but be defensive
        L1 = l1_enum[0] if l1_enum else "Other"
    if L2 not in intent_tree.get(L1, []):
        # Repair to Other
        if "Other" in intent_tree.get(L1, []):
            L2 = "Other"

    if Sent not in SENTIMENT_ENUM:
        Sent = "Neutral"

    return L1, L2, Sent


# ---------------------------
# Intent tree upload parsing
# ---------------------------

def parse_uploaded_intent_tree(file_bytes: bytes, filename: str) -> Dict[str, List[str]]:
    """
    Accepts:
      - CSV with columns L1, L2
      - XLSX with columns L1, L2 (first sheet)
      - JSON:
          {"intent_tree":[{"L1":"...","L2":["..."]},...]}
          OR {"pairs":[{"L1":"...","L2":"..."}...]}
          OR {"tree":{"L1":[...]}}

    Returns dict {L1:[L2,...]}
    """
    name = filename.lower()

    if name.endswith(".csv"):
        df = pd.read_csv(BytesIO(file_bytes))
        return _tree_from_pairs_df(df)

    if name.endswith(".xlsx") or name.endswith(".xls"):
        df = pd.read_excel(BytesIO(file_bytes))
        return _tree_from_pairs_df(df)

    if name.endswith(".json"):
        obj = json.loads(file_bytes.decode("utf-8"))
        return _tree_from_json(obj)

    raise ValueError("Unsupported file type. Upload CSV, XLSX, or JSON.")


def _tree_from_pairs_df(df: pd.DataFrame) -> Dict[str, List[str]]:
    cols = {c.lower(): c for c in df.columns}
    if "l1" not in cols or "l2" not in cols:
        raise ValueError("Uploaded file must contain columns: L1 and L2")

    l1c, l2c = cols["l1"], cols["l2"]
    tree: Dict[str, List[str]] = {}
    for _, r in df.iterrows():
        l1 = str(r[l1c]).strip()
        l2 = str(r[l2c]).strip()
        if not l1 or not l2:
            continue
        tree.setdefault(l1, [])
        if l2 not in tree[l1]:
            tree[l1].append(l2)

    # Ensure Other exists under each L1 (optional but practical)
    for l1 in list(tree.keys()):
        if "Other" not in tree[l1]:
            tree[l1].append("Other")
        tree[l1] = sorted(tree[l1])

    return tree


def _tree_from_json(obj: dict) -> Dict[str, List[str]]:
    if isinstance(obj, dict) and "tree" in obj and isinstance(obj["tree"], dict):
        tree = {str(k).strip(): [str(x).strip() for x in v] for k, v in obj["tree"].items()}
    elif isinstance(obj, dict) and "intent_tree" in obj:
        tree = {}
        for node in obj["intent_tree"]:
            l1 = str(node.get("L1", "")).strip()
            l2s = node.get("L2", [])
            if l1 and isinstance(l2s, list):
                tree[l1] = [str(x).strip() for x in l2s if str(x).strip()]
    elif isinstance(obj, dict) and "pairs" in obj:
        # pairs = [{"L1":"...","L2":"..."}]
        df = pd.DataFrame(obj["pairs"])
        return _tree_from_pairs_df(df)
    else:
        raise ValueError("JSON format not recognized. Use tree/intent_tree/pairs structure.")

    for l1 in list(tree.keys()):
        if "Other" not in tree[l1]:
            tree[l1].append("Other")
        tree[l1] = sorted(list(dict.fromkeys(tree[l1])))

    return tree


# ---------------------------
# Excel export
# ---------------------------

def build_mapping_excel(
    mappings_df: pd.DataFrame,
    include_frequency_sheet: bool = False,
) -> bytes:
    """
    Sheet1: Mappings with headers:
      S No., Transcript text in one cell, Total AHT, L1, L2, Sentiment
    Sheet2 (optional): Intent frequency
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mappings"

    headers = ["S No.", "Transcript text in one cell", "Total AHT", "L1", "L2", "Sentiment"]
    ws.append(headers)

    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = Alignment(vertical="center", wrap_text=True)

    # Rows
    for _, r in mappings_df.iterrows():
        ws.append(
            [
                int(r["S No."]),
                str(r["Transcript text in one cell"]),
                float(r["Total AHT"]),
                str(r["L1"]),
                str(r["L2"]),
                str(r["Sentiment"]),
            ]
        )

    # Formatting
    ws.freeze_panes = "A2"
    wrap_cols = [2]  # transcript cell
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.col_idx in wrap_cols))

    # Column widths (rough)
    col_widths = {
        1: 7,   # S No.
        2: 80,  # Transcript
        3: 12,  # AHT
        4: 18,  # L1
        5: 28,  # L2
        6: 12,  # Sentiment
    }
    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    if include_frequency_sheet:
        wf = wb.create_sheet("Intent_Frequency")
        wf_headers = ["L1", "L2", "Count", "Percent"]
        wf.append(wf_headers)
        for c in range(1, len(wf_headers) + 1):
            wf.cell(row=1, column=c).font = header_font

        freq = (
            mappings_df.groupby(["L1", "L2"], as_index=False)
            .size()
            .rename(columns={"size": "Count"})
        )
        total = float(freq["Count"].sum()) if len(freq) else 1.0
        freq["Percent"] = (freq["Count"] / total).round(4)

        for _, r in freq.sort_values(["L1", "Count"], ascending=[True, False]).iterrows():
            wf.append([r["L1"], r["L2"], int(r["Count"]), float(r["Percent"])])

        wf.freeze_panes = "A2"
        wf.column_dimensions["A"].width = 22
        wf.column_dimensions["B"].width = 34
        wf.column_dimensions["C"].width = 10
        wf.column_dimensions["D"].width = 12

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
